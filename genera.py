#!/usr/bin/env python3
"""
genera.py — Genera index.html e torneo_programma.xlsx dal file dati_torneo.json

Uso:
    python3 genera.py                  # genera entrambi i file nella cartella corrente
    python3 genera.py --only-html      # solo HTML
    python3 genera.py --only-xlsx      # solo Excel

Richiede: openpyxl, Pillow (per rigenerare loghi da immagini sorgente se necessario)
         pip install openpyxl Pillow

Input:
    - dati_torneo.json  (dati del torneo: squadre, partite, risultati)
    - logos_b64.json    (loghi in base64, già preprocessati)

Output:
    - index.html
    - torneo_programma.xlsx
"""

import json
import sys
import argparse
from pathlib import Path
from collections import defaultdict


# ═══════════════════════════════════════════════════════════════════════════
# LOAD DATA
# ═══════════════════════════════════════════════════════════════════════════

def load_data(base_dir='.'):
    """Carica i dati del torneo e i loghi."""
    base = Path(base_dir)
    with open(base / 'dati_torneo.json', encoding='utf-8') as f:
        dati = json.load(f)
    with open(base / 'logos_b64.json', encoding='utf-8') as f:
        loghi = json.load(f)
    return dati, loghi


def get_logo(nome_breve, dati, loghi):
    """Restituisce il data URI del logo per il nome breve squadra."""
    for girone, squadre in dati['squadre'].items():
        for sq in squadre:
            if sq['nome_breve'] == nome_breve:
                return loghi.get(sq['logo_key'], '')
    # Logo generico (trofeo, shoemakers)
    key_map = {'trofeo': 'trofeo', 'shoemakers_societa': 'shoemakers'}
    if nome_breve in key_map:
        return loghi.get(key_map[nome_breve], '')
    return ''


def calcola_classifica(dati, girone_code):
    """
    Calcola la classifica di un girone basandosi sulle partite giocate.

    Regole italiane FIP:
    - Vittoria = 2 punti
    - Sconfitta = 1 punto (anche dopo tempi supplementari)
    - Sconfitta a tavolino = 0 punti (non gestito qui)

    Ritorna una lista ordinata: [{nome, G, V, P, PF, PS, Pti, pos}, ...]
    Ordine: 1) Pti desc, 2) differenza punti desc, 3) PF desc, 4) ordine squadra
    """
    squadre = dati['squadre'][girone_code]
    # Inizializza
    stats = {}
    for sq in squadre:
        stats[sq['nome_breve']] = {
            'nome_breve': sq['nome_breve'],
            'nome_esteso': sq['nome_esteso'],
            'logo_key': sq['logo_key'],
            'G': 0, 'V': 0, 'P': 0, 'PF': 0, 'PS': 0, 'Pti': 0
        }

    # Scorri le partite del girone
    for p in dati['partite']:
        if p.get('placeholder'):
            continue
        if p['fase'] != f"Girone {girone_code}":
            continue
        if not p.get('giocata'):
            continue

        s1, s2 = p['squadra1'], p['squadra2']
        pt1, pt2 = p['punti1'], p['punti2']

        if s1 not in stats or s2 not in stats:
            continue

        stats[s1]['G'] += 1
        stats[s2]['G'] += 1
        stats[s1]['PF'] += pt1
        stats[s1]['PS'] += pt2
        stats[s2]['PF'] += pt2
        stats[s2]['PS'] += pt1

        if pt1 > pt2:
            stats[s1]['V'] += 1
            stats[s2]['P'] += 1
            stats[s1]['Pti'] += 2
            stats[s2]['Pti'] += 1
        elif pt2 > pt1:
            stats[s2]['V'] += 1
            stats[s1]['P'] += 1
            stats[s2]['Pti'] += 2
            stats[s1]['Pti'] += 1
        # Pareggio non ammesso nel basket; se per errore 0-0 non giocata, già filtrata

    # Ordina: mantieni ordine squadre quando nessuna partita giocata
    classifica = list(stats.values())
    # Se nessuna partita giocata (torneo non iniziato): nessuna posizione
    qualche_partita = any(s['G'] > 0 for s in classifica)

    if qualche_partita:
        classifica.sort(key=lambda x: (
            -x['Pti'],
            -(x['PF'] - x['PS']),
            -x['PF']
        ))
        for i, s in enumerate(classifica, 1):
            s['pos'] = str(i)
    else:
        for s in classifica:
            s['pos'] = '–'

    return classifica


# ═══════════════════════════════════════════════════════════════════════════
# HTML GENERATION
# ═══════════════════════════════════════════════════════════════════════════

def genera_html(dati, loghi, output_path='index.html'):
    """Genera l'HTML standalone completo."""
    anno = dati['torneo']['anno']
    giorni = dati['giorni']

    # Aggiungi cerimonie alla lista partite (per visualizzazione)
    partite_e_cerimonie = list(dati['partite'])
    for cer in dati['cerimonie']:
        partite_e_cerimonie.append({
            'giorno': cer['giorno'],
            'ora': cer['ora_inizio'],
            'ora_fine': cer['ora_fine'],
            'impianto': cer['impianto'],
            'categoria': 'TUTTI',
            'squadra1': cer['titolo'],
            'squadra2': '',
            'fase': cer['descrizione'],
            'tipo': 'cerimonia',
            'placeholder': False,
        })

    go = {'G1': 1, 'G2': 2, 'G3': 3}
    partite_e_cerimonie.sort(key=lambda x: (go[x['giorno']], x['ora'], x.get('impianto', 'A')))

    # Raggruppa per slot
    by_slot = defaultdict(list)
    for m in partite_e_cerimonie:
        by_slot[(m['giorno'], m['ora'])].append(m)

    # Impianti: mappa sigla -> nome (usa il nome esatto dal JSON, senza auto-prefix)
    imp_names = {}
    for imp in dati['impianti']:
        if imp['tipo'] != 'pool':
            imp_names[imp['sigla']] = imp['nome']

    # ── SEZIONE PROGRAMMA ──
    slots_html = ''
    match_num_display = 0
    for gcode in ['G1', 'G2', 'G3']:
        slots_html += f'<div class="day-block"><div class="day-header"><span class="day-tag">{gcode}</span><h2>{giorni[gcode]}</h2></div>\n'
        orari = sorted(set(m['ora'] for m in partite_e_cerimonie if m['giorno'] == gcode))
        for ora in orari:
            partite_slot = by_slot[(gcode, ora)]
            if not partite_slot:
                continue
            is_cer_only = all(p.get('tipo') == 'cerimonia' for p in partite_slot)
            try:
                h, mm = ora.split(':')
                end_o = f'{int(h)+2:02d}:{mm}'
            except Exception:
                end_o = ''

            # Badge mensa
            mensa = ''
            if gcode == 'G1':
                if ora == '10:00':
                    mensa = '<span class="badge-mensa verde">🍽 Mensa 13:00</span>'
                elif ora == '12:00':
                    mensa = '<span class="badge-mensa arancio">✈️ Fuori regione · Mensa 14:00</span>'
            elif gcode == 'G2':
                if ora == '09:00':
                    mensa = '<span class="badge-mensa verde">🍽 Mensa 12:00</span>'
                elif ora == '11:00':
                    mensa = '<span class="badge-mensa arancio">🍽 Mensa 13:00</span>'
            elif gcode == 'G3':
                if ora == '09:00':
                    mensa = '<span class="badge-mensa verde">🍽 Mensa 12:00</span>'
                elif ora == '11:00':
                    mensa = '<span class="badge-mensa arancio">🍽 Mensa 13:00</span>'
                elif ora == '13:00':
                    mensa = '<span class="badge-mensa gold">🏆 Finalissima U13</span>'
                elif ora == '15:00':
                    mensa = '<span class="badge-mensa gold">🏆 Finalissima U14</span>'

            end_disp = partite_slot[0].get('ora_fine', end_o) if is_cer_only else end_o
            slots_html += f'<div class="slot-block"><div class="slot-header"><div class="slot-time">{ora}</div><div class="slot-end">→ {end_disp}</div>{mensa}</div><div class="slot-matches">\n'

            for m in partite_slot:
                is_ceremony = m.get('tipo') == 'cerimonia'
                is_pala = m['impianto'] == 'A'
                fu = m['fase'].upper()
                is_fin = 'FINAL' in fu and not is_ceremony and '🏆' in m['fase']
                is_semi = 'SEMI' in fu

                cls = 'mc'
                if is_ceremony:
                    cls += ' mc-cer'
                elif is_fin:
                    cls += ' mc-fin'
                elif is_semi:
                    cls += ' mc-semi'
                elif is_pala:
                    cls += ' mc-pala'

                if is_ceremony:
                    slots_html += f'<div class="{cls}"><div class="mc-top"><span class="mc-imp mc-imp-pala">🏟 PalaCardelli</span><span class="mc-time-cer">{m["ora"]}–{m.get("ora_fine","")}</span></div><div class="mc-cer-title">🎉 {m["squadra1"]}</div><div class="mc-cer-desc">{m["fase"]}</div></div>\n'
                else:
                    match_num_display += 1
                    t1, t2 = m['squadra1'], m['squadra2']
                    is_placeholder = m.get('placeholder', False)
                    l1 = get_logo(t1, dati, loghi) if not is_placeholder else ''
                    l2 = get_logo(t2, dati, loghi) if not is_placeholder else ''

                    # Score display
                    if m.get('giocata'):
                        sc1 = str(m['punti1'])
                        sc2 = str(m['punti2'])
                    elif is_placeholder:
                        sc1 = sc2 = '–'
                    else:
                        sc1 = sc2 = '0'

                    cat_cls = 'cu13' if m['categoria'] == 'U13' else 'cu14'
                    imp_cls = 'mc-imp-pala' if is_pala else ''

                    if l1 and l2:
                        body = (f'<div class="mc-team"><img src="{l1}" alt="{t1}"><span>{t1}</span><span class="mc-score">{sc1}</span></div>'
                                f'<div class="mc-vs">vs</div>'
                                f'<div class="mc-team"><img src="{l2}" alt="{t2}"><span>{t2}</span><span class="mc-score">{sc2}</span></div>')
                    else:
                        body = (f'<div class="mc-team mc-placeholder"><span>{t1}</span><span class="mc-score">{sc1}</span></div>'
                                f'<div class="mc-vs">vs</div>'
                                f'<div class="mc-team mc-placeholder"><span>{t2}</span><span class="mc-score">{sc2}</span></div>')
                    slots_html += f'<div class="{cls}"><div class="mc-top"><span class="mc-imp {imp_cls}">{imp_names.get(m["impianto"], m["impianto"])}</span><span class="mc-cat {cat_cls}">{m["categoria"]}</span><span class="mc-num">#{match_num_display}</span></div><div class="mc-body">{body}</div><div class="mc-phase">{m["fase"]}</div></div>\n'
            slots_html += '</div></div>\n'
        slots_html += '</div>\n'

    # ── SEZIONE GIRONI ──
    gironi_html = ''
    for cat, gironi_cat in [('U13', ['A13', 'B13']), ('U14', ['A14', 'B14'])]:
        for gc in gironi_cat:
            squadre_g = dati['squadre'][gc]
            ccls = 'u13' if cat == 'U13' else 'u14'
            cards = ''
            for sq in squadre_g:
                logo = loghi.get(sq['logo_key'], '')
                fr_b = '<span class="fr-badge">✈ Fuori regione</span>' if sq['fuori_regione'] else ''
                cards += f'<div class="team-card {ccls}"><img src="{logo}" alt="{sq["nome_breve"]}" class="team-logo"><div class="team-info"><div class="team-slot">{sq["slot"]}</div><div class="team-name">{sq["nome_esteso"]}</div><div class="team-loc">📍 {sq["citta"]} ({sq["prov"]}) · {sq["regione"]}</div>{fr_b}</div></div>'

            # Classifica
            classifica = calcola_classifica(dati, gc)
            rows = ''
            for s in classifica:
                logo = loghi.get(s['logo_key'], '')
                diff = s['PF'] - s['PS']
                diff_str = f"{s['PF']}–{s['PS']}"
                rows += f'''<tr>
<td class="pos">{s['pos']}</td>
<td class="team"><img src="{logo}" alt="{s['nome_breve']}"><span>{s['nome_esteso']}</span></td>
<td class="num">{s['G']}</td>
<td class="num">{s['V']}</td>
<td class="num">{s['P']}</td>
<td class="num">{diff_str}</td>
<td class="pts">{s['Pti']}</td>
</tr>'''
            standings_html = f'''<div class="standings-box">
<table class="standings-table">
<thead>
<tr><th class="num">#</th><th>Squadra</th><th class="num" title="Giocate">G</th><th class="num" title="Vinte">V</th><th class="num" title="Perse">P</th><th class="num" title="Punti fatti - subiti">PF–PS</th><th class="num" title="Punti classifica">Pti</th></tr>
</thead>
<tbody>{rows}</tbody>
</table>
</div>'''

            gironi_html += f'<section class="girone-box"><div class="girone-hdr"><span class="girone-cat cat-{ccls}">{cat}</span><h3>Girone {gc}</h3><span class="girone-n">4 squadre</span></div><div class="teams-row">{cards}</div>{standings_html}</section>'

    # ── TEAM STRIP ──
    strip_nomi = []
    for gcode in ['A13', 'B13', 'A14', 'B14']:
        for sq in dati['squadre'][gcode]:
            if sq['nome_breve'] not in strip_nomi:
                strip_nomi.append(sq['nome_breve'])
    team_strip = '\n'.join(f'<img src="{loghi.get(next((s["logo_key"] for g in dati["squadre"].values() for s in g if s["nome_breve"] == t), ""), "")}" alt="{t}" title="{t}">' for t in strip_nomi)

    # ── SEZIONE MENSA ──
    # Costruisco dinamicamente dai dati
    def mensa_slot_items(slot_giorno, slot_ora):
        """Restituisce [(cat, nome_squadra, provenienza), ...] per il turno mensa corrispondente."""
        partite_slot = [p for p in dati['partite']
                        if p['giorno'] == slot_giorno and p['ora'] == slot_ora
                        and not p.get('placeholder')]
        imp_label = {'A': '🏟 PalaCardelli', 'B': 'Cintolese', 'C': 'Salutati', 'D': 'Geodetica'}
        items = []
        for p in partite_slot:
            items.append((p['categoria'], p['squadra1'], f"{imp_label[p['impianto']]} {slot_ora}"))
            items.append((p['categoria'], p['squadra2'], f"{imp_label[p['impianto']]} {slot_ora}"))
        return items

    def mensa_list_html(items):
        h = ''
        for cat, sq, da in items:
            cat_cls = 'cu13' if cat == 'U13' else 'cu14'
            logo = get_logo(sq, dati, loghi)
            h += f'<li><img src="{logo}"><span class="tl-cat {cat_cls}">{cat}</span><span class="tl-name">{sq}</span><span class="tl-from">{da}</span></li>\n'
        return h

    g1_t1 = mensa_slot_items('G1', '10:00')
    g1_t2 = mensa_slot_items('G1', '12:00')
    g2_t1 = mensa_slot_items('G2', '09:00')
    g2_t2 = mensa_slot_items('G2', '11:00')

    mensa_html = f'''<div class="info-box warn">
<h3>⚠️ Schema in revisione (4 turni × 45 min)</h3>
<p>Questa pagina mostra lo schema <strong>a 2 turni</strong> attuale. È in corso la pianificazione del passaggio a <strong>4 turni da 45 minuti</strong>.</p>
</div>

<div class="mensa-block">
<div class="mensa-day-title">GIORNO 1 — Venerdì 1 Maggio</div>
<div class="mensa-turni">
<div class="turno-box">
<div class="turno-hdr verde">🟢 Turno 13:00 · {len(g1_t1)} squadre</div>
<ul class="turno-list">{mensa_list_html(g1_t1)}</ul>
</div>
<div class="turno-box">
<div class="turno-hdr arancio">🟠 Turno 14:00 · ✈ fuori regione</div>
<ul class="turno-list">{mensa_list_html(g1_t2)}</ul>
</div>
</div>
</div>

<div class="mensa-block">
<div class="mensa-day-title">GIORNO 2 — Sabato 2 Maggio</div>
<div class="mensa-turni">
<div class="turno-box">
<div class="turno-hdr verde">🟢 Turno 12:00 · {len(g2_t1)} squadre</div>
<ul class="turno-list">{mensa_list_html(g2_t1)}</ul>
</div>
<div class="turno-box">
<div class="turno-hdr arancio">🟠 Turno 13:00 · {len(g2_t2)} squadre</div>
<ul class="turno-list">{mensa_list_html(g2_t2)}</ul>
</div>
</div>
</div>

<div class="mensa-block">
<div class="mensa-day-title">GIORNO 3 — Domenica 3 Maggio</div>
<div class="mensa-turni">
<div class="turno-box">
<div class="turno-hdr verde">🟢 Turno 12:00 · finali mattutine U13</div>
<ul class="turno-list">
<li><span class="tl-cat cu13">U13</span><span class="tl-name">Finale 3°/4°</span><span class="tl-from">Pala 09:00</span></li>
<li><span class="tl-cat cu13">U13</span><span class="tl-name">Finale 5°/6°</span><span class="tl-from">Geodetica 09:00</span></li>
<li><span class="tl-cat cu13">U13</span><span class="tl-name">Finale 7°/8°</span><span class="tl-from">Cintolese 09:00</span></li>
</ul>
</div>
<div class="turno-box">
<div class="turno-hdr arancio">🟠 Turno 13:00 · prima delle finalissime</div>
<ul class="turno-list">
<li><span class="tl-cat cu14">U14</span><span class="tl-name">Finale 3°/4°</span><span class="tl-from">Pala 11:00</span></li>
<li><span class="tl-cat cu14">U14</span><span class="tl-name">Finale 5°/6°</span><span class="tl-from">Geodetica 11:00</span></li>
<li><span class="tl-cat cu14">U14</span><span class="tl-name">Finale 7°/8°</span><span class="tl-from">Cintolese 11:00</span></li>
<li><span class="tl-cat cu13">U13</span><span class="tl-name">🏆 Finaliste U13</span><span class="tl-from">(prima della finale)</span></li>
</ul>
</div>
<div class="turno-box">
<div class="turno-hdr gold">🟡 Turno 14:30 · finaliste U14</div>
<ul class="turno-list">
<li><span class="tl-cat cu14">U14</span><span class="tl-name">🏆 Finaliste U14</span><span class="tl-from">(prima della finale 15:00)</span></li>
</ul>
</div>
</div>
</div>'''

    # ── SEZIONE IMPIANTI ──
    imp_cards = ''
    for imp in dati['impianti']:
        if imp['tipo'] == 'palazzetto':
            card_cls = 'palazzetto'
            code_style = 'color:var(--gold)'
            if imp['sigla'] == 'D':
                code_style = 'color:#fdba74'
        elif imp['tipo'] == 'pool':
            card_cls = 'pool'
            code_style = 'color:#86efac'
        else:
            card_cls = ''
            colors = {'B': '#fca5a5', 'C': '#86efac'}
            code_style = f'color:{colors.get(imp["sigla"], "var(--gold)")}'

        sigla_disp = imp['sigla']
        if imp['tipo'] == 'palazzetto' and imp['sigla'] in ['A', 'D']:
            sigla_disp = f'🏟 {imp["sigla"]}'

        desc = imp['descrizione'].replace('\n', ' ')
        imp_cards += f'''<div class="imp-card {card_cls}">
<div class="imp-code" style="{code_style}">{sigla_disp}</div>
<div class="imp-name">{imp["nome"]}</div>
<div class="imp-city">📍 {imp["indirizzo"]}</div>
<div class="imp-desc">{desc}</div>
<a href="{imp["maps_url"]}" target="_blank" rel="noopener" class="imp-map">📍 Apri su Maps</a>
</div>
'''

    trofeo_logo = loghi.get('trofeo', '')
    shoe_logo = loghi.get('shoemakers', '')

    html = f'''<!DOCTYPE html>
<html lang="it"><head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0,maximum-scale=5,viewport-fit=cover">
<meta name="theme-color" content="#0a1428">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
<meta name="format-detection" content="telephone=no">
<title>Trofeo Monsummano {anno}</title>
<style>
{_CSS}
</style></head>
<body>

<header class="hero">
<div class="hero-logos">
<img src="{trofeo_logo}" alt="Trofeo">
<img src="{shoe_logo}" alt="Shoemakers" class="shmk">
</div>
<h1>TROFEO <span>MONSUMMANO</span></h1>
<div class="hero-sub">{dati['torneo']['date']} · Under 13 &amp; Under 14</div>
<div class="hero-stats">
<span class="stat-pill"><b>16</b> squadre</span>
<span class="stat-pill"><b>40</b> partite</span>
<span class="stat-pill"><b>4</b> impianti</span>
<span class="stat-pill">🎉 2 cerimonie</span>
</div>
</header>

<div class="team-strip">{team_strip}</div>

<nav class="nav-tabs">
<button class="tab active" onclick="show('programma',event)">📅 Programma</button>
<button class="tab" onclick="show('gironi',event)">🏅 Gironi</button>
<button class="tab" onclick="show('mensa',event)">🍽 Mensa</button>
<button class="tab" onclick="show('impianti',event)">🏟 Impianti</button>
</nav>

<section class="section active" id="sec-programma">
<div class="info-box">
<h3>🏆 Gran finale al PalaCardelli domenica</h3>
<ul>
<li><strong>13:00–15:00</strong> · 🏆 Finalissima <strong>U13</strong></li>
<li><strong>15:00–17:00</strong> · 🏆 Finalissima <strong>U14</strong></li>
<li><strong>17:00–17:45</strong> · 🏆 Cerimonia di chiusura e premiazioni</li>
</ul>
<p style="margin-top:6px"><strong>Apertura</strong>: Ven 1 Maggio · 15:00–15:45 al PalaCardelli</p>
<p style="margin-top:10px"><a href="torneo_programma.xlsx" download>📥 Scarica il programma completo (Excel)</a></p>
</div>
{slots_html}
</section>

<section class="section" id="sec-gironi">
<div class="info-box">
<h3>Formato</h3>
<p>Ogni squadra gioca <strong>5 partite</strong>: 3 di girone + 2 di fase finale.</p>
<ul><li>Prime 2 di ogni girone → <strong>Semifinali Oro</strong> (1°–4° posto)</li>
<li>Ultime 2 → <strong>Semifinali Bronzo</strong> (5°–8° posto)</li></ul>
</div>
{gironi_html}
</section>

<section class="section" id="sec-mensa">
{mensa_html}
</section>

<section class="section" id="sec-impianti">
<div class="info-box">
<h3>🏟 Impianti e punto di ristoro</h3>
<p><strong>PalaCardelli</strong>, <strong>Geodetica</strong> e <strong>Pool Bar</strong> sono tutti in <strong>Piazza Sandro Pertini</strong> a Monsummano Terme — stessa posizione, massima comodità.</p>
</div>
<div class="impianti-grid">
{imp_cards}
</div>
</section>

<footer class="footer">
<b>{dati['torneo']['nome_esteso']} {anno}</b><br>
Organizzato da <b>{dati['torneo']['organizzatore']}</b><br>
{dati['torneo']['date']} · 🔵🔴
</footer>

<script>
function show(id,evt){{
document.querySelectorAll('.section').forEach(s=>s.classList.remove('active'));
document.querySelectorAll('.tab').forEach(t=>t.classList.remove('active'));
document.getElementById('sec-'+id).classList.add('active');
if(evt&&evt.currentTarget)evt.currentTarget.classList.add('active');
window.scrollTo({{top:0,behavior:'instant'}});
}}
</script>
</body></html>'''

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)

    size_kb = Path(output_path).stat().st_size / 1024
    print(f"✅ {output_path} ({size_kb:.0f} KB)")


# CSS separato per leggibilità
_CSS = '''*{box-sizing:border-box;margin:0;padding:0;-webkit-tap-highlight-color:transparent}
:root{--gold:#c9a84c;--gold-l:#e4c77a;--bg:#0a1428;--sf:#13213f;--sf2:#1a2d52;--sf3:#253c6b;--tx:#eef2f8;--dim:#a0b2cc;--mu:#6d82a1;--u13:#3b82f6;--u14:#ef4444;--cer:#a855f7}
html{-webkit-text-size-adjust:100%;scroll-behavior:smooth}
body{background:var(--bg);background-image:radial-gradient(at 20% 10%,rgba(59,130,246,.08) 0%,transparent 50%),radial-gradient(at 80% 90%,rgba(239,68,68,.08) 0%,transparent 50%);color:var(--tx);font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Arial,sans-serif;font-size:14px;line-height:1.5;min-height:100vh;padding-bottom:40px;overflow-x:hidden}
.hero{text-align:center;padding:24px 16px 20px;background:linear-gradient(180deg,rgba(10,31,92,.8) 0%,transparent 100%);border-bottom:3px solid var(--gold)}
.hero-logos{display:flex;align-items:center;justify-content:center;gap:16px;margin-bottom:14px;flex-wrap:wrap}
.hero-logos img{height:72px;width:auto;max-width:180px;object-fit:contain;filter:drop-shadow(0 6px 20px rgba(201,168,76,.3))}
.hero-logos .shmk{height:60px;width:60px;background:rgba(255,255,255,.95);padding:4px;border-radius:10px}
h1{font-family:Arial Black,sans-serif;font-size:clamp(22px,5vw,40px);font-weight:900;letter-spacing:2px;line-height:1.1;margin-bottom:4px;text-transform:uppercase}
h1 span{color:var(--gold)}
.hero-sub{font-size:12px;color:var(--dim);letter-spacing:1px;text-transform:uppercase}
.hero-stats{display:flex;justify-content:center;gap:10px;flex-wrap:wrap;margin-top:14px}
.stat-pill{background:var(--sf);border:1px solid rgba(201,168,76,.3);border-radius:20px;padding:5px 12px;font-size:11px;color:var(--dim)}
.stat-pill b{color:var(--gold);font-weight:700}
.team-strip{display:flex;gap:10px;padding:12px 14px;overflow-x:auto;-webkit-overflow-scrolling:touch;background:rgba(255,255,255,.03);border-bottom:1px solid rgba(255,255,255,.05);scrollbar-width:none}
.team-strip::-webkit-scrollbar{display:none}
.team-strip img{height:44px;width:44px;flex-shrink:0;object-fit:contain;background:rgba(255,255,255,.95);border-radius:6px;padding:3px}
.nav-tabs{position:sticky;top:0;z-index:100;background:var(--bg);padding:10px 8px;border-bottom:1px solid var(--sf3);display:flex;justify-content:center;gap:5px;overflow-x:auto;-webkit-overflow-scrolling:touch;scrollbar-width:none}
.nav-tabs::-webkit-scrollbar{display:none}
.tab{padding:8px 14px;background:var(--sf);border:1.5px solid var(--sf3);border-radius:6px;color:var(--dim);font-weight:600;font-size:11px;letter-spacing:1px;text-transform:uppercase;cursor:pointer;white-space:nowrap;font-family:inherit}
.tab.active{background:var(--sf2);border-color:var(--gold);color:var(--gold)}
.section{display:none;padding:16px 12px;max-width:1400px;margin:0 auto}
.section.active{display:block}
.day-block{margin-bottom:28px}
.day-header{display:flex;align-items:center;gap:10px;margin-bottom:14px;padding-bottom:8px;border-bottom:2px solid var(--gold)}
.day-tag{background:var(--gold);color:var(--bg);font-weight:900;padding:3px 10px;border-radius:4px;font-size:13px;letter-spacing:1px}
.day-header h2{font-size:20px;font-weight:700;letter-spacing:.5px}
.slot-block{background:var(--sf);border-radius:12px;overflow:hidden;margin-bottom:12px;border:1px solid var(--sf2)}
.slot-header{display:flex;align-items:center;gap:10px;padding:10px 14px;background:var(--sf2);border-bottom:2px solid rgba(201,168,76,.3);flex-wrap:wrap}
.slot-time{font-family:Arial Black,sans-serif;font-size:22px;font-weight:900;color:var(--gold);line-height:1}
.slot-end{font-size:11px;color:var(--mu);margin-right:auto}
.badge-mensa{font-size:10px;font-weight:700;padding:3px 9px;border-radius:10px;white-space:nowrap}
.badge-mensa.verde{background:rgba(34,197,94,.15);color:#86efac;border:1px solid rgba(34,197,94,.3)}
.badge-mensa.arancio{background:rgba(249,115,22,.15);color:#fdba74;border:1px solid rgba(249,115,22,.3)}
.badge-mensa.gold{background:rgba(201,168,76,.15);color:var(--gold-l);border:1px solid rgba(201,168,76,.3)}
.slot-matches{display:grid;grid-template-columns:repeat(auto-fill,minmax(280px,1fr));gap:1px;background:var(--sf2)}
.mc{background:var(--sf);padding:10px 12px;display:flex;flex-direction:column;gap:6px}
.mc-pala{background:linear-gradient(135deg,rgba(201,168,76,.08) 0%,var(--sf) 50%);border-left:3px solid var(--gold)}
.mc-semi{background:linear-gradient(135deg,rgba(201,168,76,.06) 0%,var(--sf) 60%);border-left:3px solid var(--gold-l)}
.mc-fin{background:linear-gradient(135deg,rgba(239,68,68,.1) 0%,var(--sf) 60%);border-left:3px solid var(--u14)}
.mc-cer{background:linear-gradient(135deg,rgba(168,85,247,.15) 0%,var(--sf) 70%);border-left:3px solid var(--cer);grid-column:1/-1}
.mc-top{display:flex;align-items:center;gap:6px;flex-wrap:wrap;font-size:10px}
.mc-imp{font-weight:700;color:var(--dim);text-transform:uppercase;letter-spacing:.5px}
.mc-imp-pala{color:var(--gold)}
.mc-cat{padding:2px 6px;border-radius:3px;font-size:9px;font-weight:900;letter-spacing:.5px}
.cu13{background:rgba(59,130,246,.2);color:#93c5fd;border:1px solid rgba(59,130,246,.4)}
.cu14{background:rgba(239,68,68,.2);color:#fca5a5;border:1px solid rgba(239,68,68,.4)}
.mc-num{margin-left:auto;color:var(--mu);font-weight:700;font-size:10px}
.mc-time-cer{margin-left:auto;font-weight:700;color:var(--cer);font-size:12px}
.mc-body{display:flex;flex-direction:column;gap:4px;margin:2px 0}
.mc-team{display:flex;align-items:center;gap:8px;font-size:13px;font-weight:600;line-height:1.3}
.mc-team img{width:32px;height:32px;flex-shrink:0;background:rgba(255,255,255,.95);border-radius:4px;padding:2px;object-fit:contain}
.mc-team.mc-placeholder{padding:4px 0;font-style:italic;color:var(--dim);font-weight:500}
.mc-team .mc-score{margin-left:auto;font-family:Arial Black,sans-serif;font-weight:900;font-size:18px;color:var(--gold);min-width:24px;text-align:right;line-height:1}
.mc-team.mc-placeholder .mc-score{color:var(--mu);opacity:.5}
.mc-vs{font-size:9px;font-weight:700;color:var(--mu);text-transform:uppercase;letter-spacing:2px;margin-left:40px;opacity:.6}
.mc-phase{font-size:10px;color:var(--mu);font-style:italic;margin-top:2px}
.mc-fin .mc-phase{color:var(--gold);font-weight:700;font-style:normal}
.mc-semi .mc-phase{color:var(--gold-l);font-weight:600;font-style:normal}
.mc-cer-title{font-size:17px;font-weight:900;color:#e9d5ff;text-align:center;margin:6px 0;letter-spacing:1px}
.mc-cer-desc{font-size:12px;color:#c4b5fd;font-style:italic;text-align:center;padding:0 8px}
.girone-box{margin-bottom:18px;background:var(--sf);border-radius:12px;overflow:hidden;border:1px solid var(--sf2)}
.girone-hdr{display:flex;align-items:center;gap:10px;padding:11px 14px;background:var(--sf2);border-bottom:2px solid var(--gold)}
.girone-cat{font-family:Arial Black,sans-serif;font-weight:900;font-size:13px;padding:3px 10px;border-radius:4px;letter-spacing:1px}
.cat-u13{background:rgba(59,130,246,.2);color:#93c5fd;border:1px solid rgba(59,130,246,.4)}
.cat-u14{background:rgba(239,68,68,.2);color:#fca5a5;border:1px solid rgba(239,68,68,.4)}
.girone-hdr h3{flex:1;font-size:18px;font-weight:700;letter-spacing:.5px}
.girone-n{font-size:11px;color:var(--mu);font-weight:600}
.teams-row{display:grid;grid-template-columns:repeat(auto-fill,minmax(300px,1fr));gap:1px;background:var(--sf2)}
.team-card{display:flex;align-items:center;gap:12px;padding:12px 14px;background:var(--sf);position:relative}
.team-card::before{content:'';position:absolute;left:0;top:8px;bottom:8px;width:3px;background:var(--gold);border-radius:0 2px 2px 0}
.team-card.u13::before{background:var(--u13)}
.team-card.u14::before{background:var(--u14)}
.team-logo{width:56px;height:56px;flex-shrink:0;border-radius:6px;object-fit:contain;background:rgba(255,255,255,.95);padding:3px}
.team-info{flex:1;min-width:0}
.team-slot{display:inline-block;background:rgba(201,168,76,.15);color:var(--gold);font-weight:900;font-size:10px;padding:2px 7px;border-radius:3px;letter-spacing:1px;margin-bottom:3px}
.team-name{font-size:14px;font-weight:700;line-height:1.2;margin-bottom:2px}
.team-loc{font-size:11px;color:var(--dim)}
.fr-badge{display:inline-block;margin-top:5px;font-size:9px;font-weight:700;padding:2px 7px;background:rgba(249,115,22,.15);color:#fdba74;border:1px solid rgba(249,115,22,.3);border-radius:10px}
.standings-box{padding:10px 14px 14px}
.standings-table{width:100%;border-collapse:collapse;font-size:13px}
.standings-table thead tr{border-bottom:2px solid var(--sf3)}
.standings-table th{text-align:left;padding:5px 8px;font-size:10px;color:var(--mu);font-weight:700;letter-spacing:.6px;text-transform:uppercase;white-space:nowrap}
.standings-table th.num{text-align:center}
.standings-table td{padding:6px 8px;border-bottom:1px solid rgba(255,255,255,.04);vertical-align:middle}
.standings-table tr:last-child td{border-bottom:none}
.standings-table td.pos{text-align:center;font-weight:900;color:var(--gold);font-size:13px;width:22px;padding-left:4px}
.standings-table td.num{text-align:center;color:var(--dim);font-size:12px;white-space:nowrap}
.standings-table td.pts{text-align:center;font-weight:900;font-size:14px;color:var(--tx)}
.standings-table td.team img{width:26px;height:26px;vertical-align:middle;border-radius:4px;object-fit:contain;background:rgba(255,255,255,.95);padding:2px;margin-right:7px;flex-shrink:0}
.standings-table td.team span{vertical-align:middle;font-weight:600;font-size:13px}
.info-box{background:var(--sf);border-left:4px solid var(--gold);border-radius:8px;padding:12px 16px;margin-bottom:18px;font-size:13px;line-height:1.6;color:var(--dim)}
.info-box strong{color:var(--tx)}
.info-box h3{font-size:14px;color:var(--gold);letter-spacing:1px;text-transform:uppercase;margin-bottom:6px}
.info-box ul{padding-left:18px;margin-top:4px}
.info-box.warn{border-left-color:#f59e0b;background:rgba(245,158,11,.06)}
.info-box.warn h3{color:#fbbf24}
.mensa-block{margin-bottom:22px}
.mensa-day-title{font-size:15px;font-weight:700;padding-bottom:6px;margin-bottom:10px;border-bottom:1px solid var(--sf3)}
.mensa-turni{display:grid;grid-template-columns:repeat(auto-fill,minmax(280px,1fr));gap:10px}
.turno-box{background:var(--sf);border-radius:10px;overflow:hidden;border:1px solid var(--sf2)}
.turno-hdr{padding:8px 12px;font-weight:700;font-size:11px;letter-spacing:.5px}
.turno-hdr.verde{background:rgba(34,197,94,.12);color:#86efac}
.turno-hdr.arancio{background:rgba(249,115,22,.12);color:#fdba74}
.turno-hdr.gold{background:rgba(201,168,76,.12);color:var(--gold-l)}
.turno-list{list-style:none;padding:0}
.turno-list li{display:flex;align-items:center;gap:8px;padding:7px 12px;font-size:12px;border-bottom:1px solid rgba(255,255,255,.04)}
.turno-list li:last-child{border-bottom:none}
.turno-list img{width:26px;height:26px;flex-shrink:0;background:rgba(255,255,255,.95);border-radius:3px;padding:2px;object-fit:contain}
.turno-list .tl-cat{font-size:9px;font-weight:900;padding:1px 5px;border-radius:3px}
.turno-list .tl-name{flex:1;font-weight:600}
.turno-list .tl-from{font-size:10px;color:var(--mu)}
.impianti-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(260px,1fr));gap:10px}
.imp-card{background:var(--sf);border-radius:10px;padding:14px;border:1px solid var(--sf2)}
.imp-card.palazzetto{background:linear-gradient(135deg,rgba(201,168,76,.1) 0%,var(--sf) 50%);border-color:rgba(201,168,76,.4)}
.imp-card.pool{background:linear-gradient(135deg,rgba(34,197,94,.1) 0%,var(--sf) 50%);border-color:rgba(34,197,94,.4)}
.imp-code{font-family:Arial Black,sans-serif;font-size:22px;color:var(--gold);margin-bottom:4px;line-height:1}
.imp-name{font-size:15px;font-weight:700;margin-bottom:4px}
.imp-city{font-size:11px;color:var(--dim);margin-bottom:6px}
.imp-desc{font-size:11px;color:var(--mu);line-height:1.5}
.imp-map{display:inline-flex;align-items:center;gap:5px;margin-top:8px;padding:5px 10px;background:rgba(59,130,246,.15);color:#93c5fd;border:1px solid rgba(59,130,246,.3);border-radius:16px;font-size:11px;font-weight:600;text-decoration:none}
.imp-card.pool .imp-map{background:rgba(34,197,94,.15);color:#86efac;border-color:rgba(34,197,94,.3)}
.footer{text-align:center;margin-top:40px;padding:20px 14px;border-top:1px solid var(--sf2);font-size:11px;color:var(--mu);line-height:1.6}
.footer b{color:var(--gold);font-weight:700}
@media (max-width:500px){body{font-size:13px}.hero{padding:18px 12px 16px}.hero-logos{gap:12px}.hero-logos img{height:56px}.hero-logos .shmk{height:48px;width:48px}h1{font-size:22px;letter-spacing:1.5px}.stat-pill{font-size:10px;padding:4px 10px}.section{padding:12px 8px}.day-header h2{font-size:16px}.slot-header{padding:8px 12px;gap:8px}.slot-time{font-size:18px}.badge-mensa{font-size:9px;padding:2px 7px}.slot-matches{grid-template-columns:1fr}.mc-team{font-size:12px}.mc-team img{width:28px;height:28px}.teams-row{grid-template-columns:1fr}.team-logo{width:48px;height:48px}.team-name{font-size:13px}.girone-hdr h3{font-size:16px}.impianti-grid,.mensa-turni{grid-template-columns:1fr}}'''


# ═══════════════════════════════════════════════════════════════════════════
# EXCEL GENERATION
# ═══════════════════════════════════════════════════════════════════════════

def genera_excel(dati, loghi, output_path='torneo_programma.xlsx'):
    """Genera il file Excel con fogli multipli."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    F_TITLE = Font(name='Calibri', size=18, bold=True, color='0A1F5C')
    F_SUB = Font(name='Calibri', size=11, italic=True, color='666666')
    F_HDR = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    F_BODY = Font(name='Calibri', size=10)
    F_BOLD = Font(name='Calibri', size=10, bold=True)
    F_GOLD = Font(name='Calibri', size=10, bold=True, color='8B6000')
    F_GREEN = Font(name='Calibri', size=10, bold=True, color='166534')
    F_LINK = Font(name='Calibri', size=10, color='1E40AF', underline='single')
    F_CER = Font(name='Calibri', size=11, bold=True, color='581C87')
    F_SCORE = Font(name='Calibri', size=12, bold=True, color='8B6000')

    FILL_HDR = PatternFill('solid', fgColor='0A1F5C')
    FILL_DAY = PatternFill('solid', fgColor='162847')
    FILL_PALA = PatternFill('solid', fgColor='FFFBF0')
    FILL_POOL = PatternFill('solid', fgColor='EFFAE8')
    FILL_U13 = PatternFill('solid', fgColor='DDEEFF')
    FILL_U14 = PatternFill('solid', fgColor='FFDDDD')
    FILL_FIN = PatternFill('solid', fgColor='FFF3CC')
    FILL_SEMI = PatternFill('solid', fgColor='FFF8E0')
    FILL_CER = PatternFill('solid', fgColor='EDE9FE')
    FILL_SCORE = PatternFill('solid', fgColor='FFF8E8')

    thin = Side(border_style='thin', color='CCCCCC')
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

    wb = Workbook()
    anno = dati['torneo']['anno']

    # ── SHEET 1: PROGRAMMA ──
    ws = wb.active
    ws.title = 'Programma'
    ws.merge_cells('A1:J1')
    ws['A1'] = f'{dati["torneo"]["nome_esteso"]} · {dati["torneo"]["date"]}'
    ws['A1'].font = F_TITLE
    ws['A1'].alignment = CENTER
    ws.row_dimensions[1].height = 32
    ws.merge_cells('A2:J2')
    ws['A2'] = '16 squadre · 40 partite · 4 impianti · Entrambe le finalissime al Palazzetto (dom 13 e 15)'
    ws['A2'].font = F_SUB
    ws['A2'].alignment = CENTER

    headers = ['#', 'Giorno', 'Data', 'Ora', 'Impianto', 'Cat.', 'Squadra 1', 'Pti', 'Squadra 2', 'Fase']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = F_HDR
        c.fill = FILL_HDR
        c.alignment = CENTER
        c.border = BORDER

    data_giorno = {'G1': 'Ven 1 Mag', 'G2': 'Sab 2 Mag', 'G3': 'Dom 3 Mag'}
    giorno_full = {'G1': 'Giorno 1', 'G2': 'Giorno 2', 'G3': 'Giorno 3'}
    # Impianti: mappa sigla -> nome (usa il nome esatto dal JSON)
    imp_names = {}
    for imp in dati['impianti']:
        if imp['tipo'] != 'pool':
            imp_names[imp['sigla']] = imp['nome']

    # Unisci partite + cerimonie, ordina
    partite_e_cer = list(dati['partite'])
    for cer in dati['cerimonie']:
        partite_e_cer.append({
            'giorno': cer['giorno'], 'ora': cer['ora_inizio'], 'ora_fine': cer['ora_fine'],
            'impianto': cer['impianto'], 'categoria': 'TUTTI',
            'squadra1': cer['titolo'], 'squadra2': '', 'fase': cer['descrizione'],
            'tipo': 'cerimonia', 'placeholder': False, 'giocata': False, 'punti1': 0, 'punti2': 0,
        })
    go = {'G1': 1, 'G2': 2, 'G3': 3}
    partite_e_cer.sort(key=lambda x: (go[x['giorno']], x['ora'], x.get('impianto', 'A')))

    row = 5
    match_num = 0
    prev_g = None
    for m in partite_e_cer:
        if m['giorno'] != prev_g:
            if prev_g is not None:
                row += 1
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
            c = ws.cell(row=row, column=1, value=f"  {giorno_full[m['giorno']]} — {data_giorno[m['giorno']]}")
            c.font = Font(name='Calibri', size=13, bold=True, color='FFFFFF')
            c.fill = FILL_DAY
            c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            ws.row_dimensions[row].height = 22
            row += 1
            prev_g = m['giorno']

        is_cer = m.get('tipo') == 'cerimonia'
        is_pala = m['impianto'] == 'A'
        fu = m['fase'].upper()
        is_fin = 'FINAL' in fu and not is_cer and 'CHIUSURA' not in fu and 'APERTURA' not in fu
        is_semi = 'SEMI' in fu

        if is_cer:
            n_disp = ''
            sq1_text = m['squadra1']
            punti_text = ''
            sq2_text = ''
            ora_disp = f"{m['ora']}–{m.get('ora_fine','')}"
        else:
            match_num += 1
            n_disp = match_num
            sq1_text = m['squadra1']
            sq2_text = m['squadra2']
            if m.get('giocata'):
                punti_text = f"{m['punti1']} – {m['punti2']}"
            elif m.get('placeholder'):
                punti_text = '–'
            else:
                punti_text = '0 – 0'
            ora_disp = m['ora']

        vals = [n_disp, giorno_full[m['giorno']], data_giorno[m['giorno']],
                ora_disp, imp_names.get(m['impianto'], m['impianto']),
                m['categoria'], sq1_text, punti_text, sq2_text, m['fase']]

        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.font = F_BODY
            c.alignment = CENTER if col not in [7, 9, 10] else LEFT
            c.border = BORDER
            if is_cer:
                c.fill = FILL_CER
                c.font = F_CER if col == 7 else Font(name='Calibri', size=10, bold=True, color='581C87')
            elif col == 5 and is_pala:
                c.fill = FILL_PALA
                c.font = F_GOLD
            elif col == 6 and not is_cer:
                c.fill = FILL_U13 if m['categoria'] == 'U13' else FILL_U14
                c.font = F_BOLD
            elif col == 8 and not is_cer:
                # Colonna Punti
                c.font = F_SCORE
                if m.get('giocata'):
                    c.fill = FILL_SCORE
                c.alignment = CENTER
            elif col == 10:
                if is_fin:
                    c.fill = FILL_FIN
                    c.font = F_GOLD
                elif is_semi:
                    c.fill = FILL_SEMI
            if col == 7 and '🏆' in m['fase'] and not is_cer:
                c.font = F_BOLD
                c.fill = FILL_FIN
        row += 1

    for col, w in enumerate([5, 10, 12, 13, 22, 8, 24, 12, 24, 28], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = 'A5'

    # ── SHEET 2: CLASSIFICHE ──
    ws_c = wb.create_sheet('Classifiche')
    ws_c.merge_cells('A1:G1')
    ws_c['A1'] = 'CLASSIFICHE GIRONI'
    ws_c['A1'].font = F_TITLE
    ws_c['A1'].alignment = CENTER
    ws_c.row_dimensions[1].height = 30

    for col, h in enumerate(['#', 'Squadra', 'G', 'V', 'P', 'PF–PS', 'Pti'], 1):
        c = ws_c.cell(row=3, column=col, value=h)
        c.font = F_HDR
        c.fill = FILL_HDR
        c.alignment = CENTER
        c.border = BORDER

    row = 4
    for cat, gironi_cat in [('U13', ['A13', 'B13']), ('U14', ['A14', 'B14'])]:
        for gc in gironi_cat:
            ws_c.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            c = ws_c.cell(row=row, column=1, value=f"  {cat} · Girone {gc}")
            c.font = Font(name='Calibri', size=13, bold=True, color='FFFFFF')
            c.fill = FILL_DAY
            c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            ws_c.row_dimensions[row].height = 22
            row += 1
            classifica = calcola_classifica(dati, gc)
            for s in classifica:
                vals = [s['pos'], s['nome_esteso'], s['G'], s['V'], s['P'],
                        f"{s['PF']}–{s['PS']}", s['Pti']]
                for col, v in enumerate(vals, 1):
                    c = ws_c.cell(row=row, column=col, value=v)
                    c.font = F_BODY
                    c.alignment = CENTER if col != 2 else LEFT
                    c.border = BORDER
                    if col == 1:
                        c.font = F_GOLD
                    elif col == 7:
                        c.font = F_GOLD
                        c.fill = FILL_FIN
                row += 1
            row += 1

    for col, w in enumerate([6, 30, 6, 6, 6, 12, 8], 1):
        ws_c.column_dimensions[get_column_letter(col)].width = w
    ws_c.freeze_panes = 'A4'

    # ── SHEET 3: GIRONI & SQUADRE ──
    ws2 = wb.create_sheet('Gironi & Squadre')
    ws2.merge_cells('A1:E1')
    ws2['A1'] = 'GIRONI & SQUADRE PARTECIPANTI'
    ws2['A1'].font = F_TITLE
    ws2['A1'].alignment = CENTER
    ws2.row_dimensions[1].height = 30

    for col, h in enumerate(['Slot', 'Squadra', 'Città', 'Prov.', 'Regione'], 1):
        c = ws2.cell(row=3, column=col, value=h)
        c.font = F_HDR
        c.fill = FILL_HDR
        c.alignment = CENTER
        c.border = BORDER

    row = 4
    for cat, gironi_cat in [('U13', ['A13', 'B13']), ('U14', ['A14', 'B14'])]:
        for gc in gironi_cat:
            ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            c = ws2.cell(row=row, column=1, value=f"  {cat} · Girone {gc}")
            c.font = Font(name='Calibri', size=13, bold=True, color='FFFFFF')
            c.fill = FILL_DAY
            c.alignment = Alignment(horizontal='left', indent=1, vertical='center')
            ws2.row_dimensions[row].height = 22
            row += 1
            for sq in dati['squadre'][gc]:
                fill_cat = FILL_U13 if cat == 'U13' else FILL_U14
                nome_disp = sq['nome_esteso'] + ('  ✈️ FUORI REGIONE' if sq['fuori_regione'] else '')
                for col, v in enumerate([sq['slot'], nome_disp, sq['citta'], sq['prov'], sq['regione']], 1):
                    c = ws2.cell(row=row, column=col, value=v)
                    c.font = F_BODY
                    c.alignment = CENTER if col != 2 else LEFT
                    c.border = BORDER
                    if col == 1:
                        c.font = F_GOLD
                        c.fill = fill_cat
                    if col == 2 and sq['fuori_regione']:
                        c.font = Font(name='Calibri', size=10, bold=True, color='B45309')
                row += 1
            row += 1

    for col, w in enumerate([10, 36, 24, 8, 18], 1):
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.freeze_panes = 'A4'

    # ── SHEET 4: IMPIANTI ──
    ws4 = wb.create_sheet('Impianti')
    ws4.merge_cells('A1:E1')
    ws4['A1'] = 'IMPIANTI & PUNTO DI RISTORO'
    ws4['A1'].font = F_TITLE
    ws4['A1'].alignment = CENTER
    ws4.row_dimensions[1].height = 30

    for col, h in enumerate(['Sigla', 'Nome', 'Indirizzo', 'Descrizione', 'Link Maps'], 1):
        c = ws4.cell(row=3, column=col, value=h)
        c.font = F_HDR
        c.fill = FILL_HDR
        c.alignment = CENTER
        c.border = BORDER

    row = 4
    for imp in dati['impianti']:
        fill = FILL_PALA if imp['tipo'] == 'palazzetto' else (FILL_POOL if imp['tipo'] == 'pool' else PatternFill())
        name_font = F_GOLD if imp['tipo'] == 'palazzetto' else (F_GREEN if imp['tipo'] == 'pool' else F_BODY)
        c1 = ws4.cell(row=row, column=1, value=imp['sigla'])
        c1.font = name_font
        c1.alignment = CENTER
        c1.fill = fill
        c1.border = BORDER
        c2 = ws4.cell(row=row, column=2, value=imp['nome'])
        c2.font = name_font
        c2.alignment = LEFT
        c2.fill = fill
        c2.border = BORDER
        c3 = ws4.cell(row=row, column=3, value=imp['indirizzo'])
        c3.font = F_BODY
        c3.alignment = LEFT
        c3.fill = fill
        c3.border = BORDER
        c4 = ws4.cell(row=row, column=4, value=imp['descrizione'])
        c4.font = F_BODY
        c4.alignment = LEFT
        c4.fill = fill
        c4.border = BORDER
        c5 = ws4.cell(row=row, column=5, value='📍 Apri su Maps')
        c5.font = F_LINK
        c5.hyperlink = imp['maps_url']
        c5.alignment = CENTER
        c5.fill = fill
        c5.border = BORDER
        ws4.row_dimensions[row].height = 48
        row += 1

    for col, w in enumerate([8, 24, 36, 52, 20], 1):
        ws4.column_dimensions[get_column_letter(col)].width = w

    # ── SHEET 5: CERIMONIE ──
    ws5 = wb.create_sheet('Cerimonie')
    ws5.merge_cells('A1:D1')
    ws5['A1'] = 'CERIMONIE UFFICIALI'
    ws5['A1'].font = F_TITLE
    ws5['A1'].alignment = CENTER

    for col, h in enumerate(['Evento', 'Giorno', 'Orario', 'Luogo'], 1):
        c = ws5.cell(row=3, column=col, value=h)
        c.font = F_HDR
        c.fill = FILL_HDR
        c.alignment = CENTER
        c.border = BORDER

    row = 4
    giorno_full_label = {'G1': 'Ven 1 Maggio', 'G2': 'Sab 2 Maggio', 'G3': 'Dom 3 Maggio'}
    for cer in dati['cerimonie']:
        emoji = '🎉' if cer['tipo'] == 'apertura' else '🏆'
        ev = f"{emoji} {cer['titolo']}"
        luogo = imp_names.get(cer['impianto'], cer['impianto'])
        ora = f"{cer['ora_inizio']} – {cer['ora_fine']}"
        for col, v in enumerate([ev, giorno_full_label[cer['giorno']], ora, luogo], 1):
            c = ws5.cell(row=row, column=col, value=v)
            c.font = F_CER if col == 1 else F_BOLD
            c.alignment = CENTER if col != 1 else LEFT
            c.fill = FILL_CER
            c.border = BORDER
        row += 1

    for col, w in enumerate([30, 16, 18, 24], 1):
        ws5.column_dimensions[get_column_letter(col)].width = w

    wb.save(output_path)
    size_kb = Path(output_path).stat().st_size / 1024
    print(f"✅ {output_path} ({size_kb:.0f} KB)")


# ═══════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description='Genera HTML e Excel dal JSON dati del torneo')
    parser.add_argument('--only-html', action='store_true', help='Genera solo HTML')
    parser.add_argument('--only-xlsx', action='store_true', help='Genera solo Excel')
    parser.add_argument('--input-dir', default='.', help='Directory con dati_torneo.json e logos_b64.json')
    parser.add_argument('--output-dir', default='.', help='Directory di output')
    args = parser.parse_args()

    dati, loghi = load_data(args.input_dir)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"📂 Caricati {len(dati['partite'])} partite, {sum(len(s) for s in dati['squadre'].values())} squadre, {len(loghi)} loghi")

    if not args.only_xlsx:
        genera_html(dati, loghi, output_dir / 'index.html')
    if not args.only_html:
        genera_excel(dati, loghi, output_dir / 'torneo_programma.xlsx')


if __name__ == '__main__':
    main()
